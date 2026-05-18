from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from inspections.models import Evaluation, StatisticalRecord


def clean(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_name(value):
    return ' '.join(clean(value).split()).casefold()


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    text = clean(value)
    if not text:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


class Command(BaseCommand):
    help = 'Import department statistics from the 2026 Excel workbook and link rows to matching reports.'

    def add_arguments(self, parser):
        parser.add_argument('workbook', type=str, help='Path to the statistics .xlsx file')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required. Install project requirements first.') from exc

        workbook_path = Path(options['workbook'])
        if not workbook_path.exists():
            raise CommandError(f'File not found: {workbook_path}')

        reports_by_name = {
            normalize_name(evaluation.facility_name): evaluation
            for evaluation in Evaluation.objects.all()
            if evaluation.facility_name
        }

        workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
        imported = 0
        linked = 0

        importers = {
            'الزيارات': self.import_visits,
            'المصانع': self.import_factories,
        }

        for worksheet in workbook.worksheets:
            sheet_name = clean(worksheet.title)
            importer = importers.get(sheet_name)
            if not importer:
                continue
            sheet_imported, sheet_linked = importer(worksheet, reports_by_name)
            imported += sheet_imported
            linked += sheet_linked

        self.stdout.write(self.style.SUCCESS(f'Imported {imported} statistical records; linked {linked} to reports.'))

    def save_record(self, worksheet, row_number, reports_by_name, **fields):
        facility_name = clean(fields.get('facility_name'))
        report = reports_by_name.get(normalize_name(facility_name)) if facility_name else None
        fields['facility_name'] = facility_name
        fields['report'] = report
        StatisticalRecord.objects.update_or_create(
            source_sheet=clean(worksheet.title),
            source_row=row_number,
            defaults=fields,
        )
        return bool(report)

    def import_visits(self, worksheet, reports_by_name):
        imported = 0
        linked = 0
        for row_number in range(11, worksheet.max_row + 1):
            values = [worksheet.cell(row=row_number, column=column).value for column in range(1, 9)]
            if not values[0] and not values[2]:
                continue
            is_linked = self.save_record(
                worksheet,
                row_number,
                reports_by_name,
                category=clean(values[1]),
                facility_name=clean(values[2]),
                activity_type=clean(values[3]),
                activity_category=clean(values[4]),
                visit_date=parse_date(values[5]),
                action_taken=clean(values[7]),
            )
            imported += 1
            linked += int(is_linked)
        return imported, linked

    def import_factories(self, worksheet, reports_by_name):
        imported = 0
        linked = 0
        for row_number in range(13, worksheet.max_row + 1):
            values = [worksheet.cell(row=row_number, column=column).value for column in range(1, 7)]
            if not values[0] and not values[1]:
                continue
            is_linked = self.save_record(
                worksheet,
                row_number,
                reports_by_name,
                category='مصنع مستوف لنظم الجودة',
                facility_name=clean(values[1]),
                activity_type=clean(values[2]),
                governorate=clean(values[3]),
                contact_info=clean(values[4]),
                quality_systems=clean(values[5]),
            )
            imported += 1
            linked += int(is_linked)
        return imported, linked
